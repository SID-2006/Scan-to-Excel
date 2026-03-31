import os
import re
import tempfile
from difflib import SequenceMatcher
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Sequence, Tuple

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from paddlex import create_pipeline

from correction_layer import run_correction_engine
from ml_text_corrector import load_model


EXPECTED_COLS = 9
EXPECTED_CLASS_ROWS = 8
COLUMN_WIDTH_RATIOS = [5, 25, 10, 10, 10, 10, 10, 10, 10]
JTE_SCHEMA = [
    "SN",
    "Volunteer/Teacher's Name",
    "In-time",
    "Out-time",
    "Class Taught",
    "No of students",
    "Subject",
    "Class Activity",
    "Homework",
]


_OCR_PIPELINE = None
_ML_PREDICTOR = None
_ML_TRIED = False


def _get_ocr_pipeline():
    global _OCR_PIPELINE
    if _OCR_PIPELINE is None:
        _OCR_PIPELINE = create_pipeline(pipeline="OCR")
    return _OCR_PIPELINE


def _get_ml_predictor():
    global _ML_PREDICTOR, _ML_TRIED
    if _ML_TRIED:
        return _ML_PREDICTOR
    model_path = os.path.join(os.path.dirname(__file__), "models", "ocr_text_corrector.pkl")
    _ML_PREDICTOR = load_model(model_path)
    _ML_TRIED = True
    return _ML_PREDICTOR


@dataclass
class GridFixMeta:
    grid_repair_failed: bool
    invalid_row_indices: List[int]
    confidence_penalty_by_row: Dict[int, float]
    row_status_by_row: Dict[int, str]


def _cluster_positions(values: Sequence[int], tolerance: int = 10) -> List[int]:
    if not values:
        return []
    values = sorted(int(v) for v in values)
    groups: List[List[int]] = [[values[0]]]
    for v in values[1:]:
        if abs(v - groups[-1][-1]) <= tolerance:
            groups[-1].append(v)
        else:
            groups.append([v])
    return [int(round(sum(group) / len(group))) for group in groups]


def _deskew(gray: np.ndarray) -> np.ndarray:
    thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
    coords = np.column_stack(np.where(thresh > 0))
    if len(coords) < 100:
        return gray
    angle = cv2.minAreaRect(coords)[-1]
    if angle < -45:
        angle = -(90 + angle)
    else:
        angle = -angle
    h, w = gray.shape[:2]
    matrix = cv2.getRotationMatrix2D((w // 2, h // 2), angle, 1.0)
    return cv2.warpAffine(gray, matrix, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)


def preprocess_image(image: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (5, 5), 0)
    deskewed = _deskew(blur)
    binary = cv2.adaptiveThreshold(
        deskewed,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV,
        21,
        9,
    )
    close_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
    return cv2.morphologyEx(binary, cv2.MORPH_CLOSE, close_kernel, iterations=1)


def isolate_class_details_table(image: np.ndarray) -> Optional[Tuple[np.ndarray, Tuple[int, int, int, int]]]:
    """
    Isolate the CLASS DETAILS grid region using contour-based detection.
    This intentionally avoids processing the full page for OCR.
    """
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    thresh = cv2.adaptiveThreshold(
        gray,
        255,
        cv2.ADAPTIVE_THRESH_MEAN_C,
        cv2.THRESH_BINARY_INV,
        15,
        2,
    )
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return None

    h, w = image.shape[:2]
    img_area = float(h * w)
    best_bbox = None
    best_score = -1.0

    for cnt in contours:
        x, y, bw, bh = cv2.boundingRect(cnt)
        area = bw * bh
        if area < img_area * 0.10:
            continue
        if bw < w * 0.65 or bh < h * 0.22:
            continue
        if y < h * 0.20:
            # Ignore top header region candidates.
            continue
        aspect = bw / float(max(bh, 1))
        if aspect < 1.8:
            continue

        # Prefer larger lower-mid rectangular tables.
        y_center = y + (bh / 2.0)
        vertical_bias = 1.0 - abs((y_center / float(h)) - 0.62)
        score = (area / img_area) * 2.0 + max(0.0, vertical_bias)
        if score > best_score:
            best_score = score
            best_bbox = (x, y, x + bw, y + bh)

    if best_bbox is None:
        # Fallback 1: use line-density contours.
        binary = preprocess_image(image)
        horizontal, vertical = detect_table_lines(binary)
        line_mask = cv2.bitwise_or(horizontal, vertical)
        contours2, _ = cv2.findContours(line_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        best2 = None
        score2 = -1.0
        for cnt in contours2:
            x, y, bw, bh = cv2.boundingRect(cnt)
            area = bw * bh
            if area < img_area * 0.06:
                continue
            if bw < w * 0.55 or bh < h * 0.18:
                continue
            if y < h * 0.18:
                continue
            aspect = bw / float(max(bh, 1))
            if aspect < 1.7:
                continue
            s = (area / img_area) + (y / float(max(h, 1)))
            if s > score2:
                score2 = s
                best2 = (x, y, x + bw, y + bh)
        best_bbox = best2

    if best_bbox is None:
        # Fallback 2: largest contour crop then keep lower-middle strip where class grid usually is.
        largest = max(contours, key=cv2.contourArea)
        x, y, bw, bh = cv2.boundingRect(largest)
        y_start = y + int(0.34 * bh)
        y_end = y + int(0.84 * bh)
        best_bbox = (x, max(0, y_start), x + bw, min(h - 1, y_end))

    x1, y1, x2, y2 = best_bbox
    table_img = image[y1:y2, x1:x2]
    if table_img.size == 0:
        return None
    return table_img, best_bbox


def detect_table_lines(binary: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
    h, w = binary.shape
    h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(25, w // 15), 1))
    v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(25, h // 15)))
    horizontal = cv2.morphologyEx(binary, cv2.MORPH_OPEN, h_kernel, iterations=1)
    vertical = cv2.morphologyEx(binary, cv2.MORPH_OPEN, v_kernel, iterations=1)
    horizontal = cv2.morphologyEx(horizontal, cv2.MORPH_CLOSE, cv2.getStructuringElement(cv2.MORPH_RECT, (13, 1)))
    vertical = cv2.morphologyEx(vertical, cv2.MORPH_CLOSE, cv2.getStructuringElement(cv2.MORPH_RECT, (1, 13)))
    return horizontal, vertical


def _table_bbox_from_lines(horizontal: np.ndarray, vertical: np.ndarray) -> Optional[Tuple[int, int, int, int]]:
    grid = cv2.bitwise_or(horizontal, vertical)
    contours, _ = cv2.findContours(grid, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return None
    best = max(contours, key=cv2.contourArea)
    x, y, w, h = cv2.boundingRect(best)
    if w < 80 or h < 80:
        return None
    return (x, y, x + w, y + h)


def _extract_intersections(horizontal: np.ndarray, vertical: np.ndarray) -> List[Tuple[int, int]]:
    intersections = cv2.bitwise_and(horizontal, vertical)
    num_labels, _, stats, centroids = cv2.connectedComponentsWithStats(intersections, connectivity=8)
    points: List[Tuple[int, int]] = []
    for idx in range(1, num_labels):
        area = int(stats[idx, cv2.CC_STAT_AREA])
        if area < 2:
            continue
        x, y = centroids[idx]
        points.append((int(round(x)), int(round(y))))
    return points


def _expected_boundaries(left: int, right: int, ratios: Sequence[int], expected_cols: int) -> List[int]:
    total = float(sum(ratios))
    width = max(1, right - left)
    boundaries = [left]
    acc = 0.0
    for ratio in ratios:
        acc += ratio / total
        boundaries.append(left + int(round(acc * width)))
    boundaries = boundaries[: expected_cols + 1]
    boundaries[-1] = right
    for idx in range(1, len(boundaries)):
        if boundaries[idx] <= boundaries[idx - 1]:
            boundaries[idx] = boundaries[idx - 1] + 1
    return boundaries


def _align_column_boundaries(x_lines: List[int], left: int, right: int) -> List[int]:
    expected = _expected_boundaries(left, right, COLUMN_WIDTH_RATIOS, EXPECTED_COLS)
    if not x_lines:
        return expected
    x_lines = sorted(x_lines)
    selected = []
    used = set()
    for xp in expected:
        best_idx = None
        best_dist = 10**9
        for idx, xv in enumerate(x_lines):
            if idx in used:
                continue
            dist = abs(xv - xp)
            if dist < best_dist:
                best_dist = dist
                best_idx = idx
        if best_idx is not None and best_dist <= max(10, (right - left) // 25):
            selected.append(x_lines[best_idx])
            used.add(best_idx)
        else:
            selected.append(xp)
    selected = sorted(selected[: EXPECTED_COLS + 1])
    for i in range(1, len(selected)):
        if selected[i] <= selected[i - 1]:
            selected[i] = selected[i - 1] + 1
    selected[0] = left
    selected[-1] = right
    return selected


def _valid_rows_from_y_lines(y_lines: List[int], min_height: int = 16) -> List[int]:
    if len(y_lines) < 2:
        return y_lines
    out = [y_lines[0]]
    for y in y_lines[1:]:
        if y - out[-1] >= min_height:
            out.append(y)
    if out[-1] != y_lines[-1]:
        out.append(y_lines[-1])
    return out


def _enforce_expected_row_boundaries(
    y_lines: List[int], top: int, bottom: int, expected_rows: int = EXPECTED_CLASS_ROWS + 1
) -> List[int]:
    """
    Enforce approximately expected row count (header + 8 data rows).
    Returns row boundaries (count = expected_rows + 1).
    """
    target_boundaries = expected_rows + 1
    if top >= bottom:
        return [0, 1]
    expected = [top + int(round((bottom - top) * i / float(expected_rows))) for i in range(target_boundaries)]
    if len(y_lines) < 2:
        return expected
    y_lines = sorted(y_lines)

    selected = []
    for yp in expected:
        nearest = min(y_lines, key=lambda v: abs(v - yp))
        selected.append(nearest)
    selected[0] = top
    selected[-1] = bottom
    selected = sorted(selected)
    for idx in range(1, len(selected)):
        if selected[idx] <= selected[idx - 1]:
            selected[idx] = selected[idx - 1] + 1
    return selected


def validate_and_fix_grid(
    points: List[Tuple[int, int]],
    table_bbox: Tuple[int, int, int, int],
) -> Tuple[List[int], List[int], GridFixMeta]:
    x1, y1, x2, y2 = table_bbox
    x_lines = _cluster_positions([p[0] for p in points], tolerance=12)
    y_lines = _cluster_positions([p[1] for p in points], tolerance=12)
    y_lines = [y for y in y_lines if y1 <= y <= y2]
    x_lines = [x for x in x_lines if x1 <= x <= x2]

    aligned_x = _align_column_boundaries(x_lines, x1, x2)
    fixed_y = _valid_rows_from_y_lines(y_lines, min_height=18)
    fixed_y = _enforce_expected_row_boundaries(fixed_y, y1, y2, expected_rows=EXPECTED_CLASS_ROWS + 1)

    invalid_rows = []
    penalties = {}
    status = {}
    for idx in range(max(0, len(fixed_y) - 1)):
        h = fixed_y[idx + 1] - fixed_y[idx]
        if h < 14:
            invalid_rows.append(idx)
            penalties[idx] = 0.25
            status[idx] = "repaired"
        else:
            penalties[idx] = 0.0
            status[idx] = "ok"

    grid_failed = len(aligned_x) != EXPECTED_COLS + 1 or len(fixed_y) < 2
    return aligned_x, fixed_y, GridFixMeta(
        grid_repair_failed=grid_failed,
        invalid_row_indices=invalid_rows,
        confidence_penalty_by_row=penalties,
        row_status_by_row=status,
    )


def _ocr_crop_with_conf(crop: np.ndarray) -> Tuple[str, float]:
    if crop is None or crop.size == 0:
        return "", 0.0
    if crop.shape[0] < 4 or crop.shape[1] < 4:
        return "", 0.0
    # Resize before OCR for better handwritten legibility.
    crop = cv2.resize(crop, None, fx=2.0, fy=2.0, interpolation=cv2.INTER_CUBIC)
    ocr = _get_ocr_pipeline()
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        cv2.imwrite(tmp.name, crop)
        path = tmp.name
    text_parts = []
    confs = []
    try:
        preds = list(ocr.predict(path))
        for pred in preds:
            rec_texts = getattr(pred, "rec_texts", None)
            rec_scores = getattr(pred, "rec_scores", None)
            if rec_texts is None and isinstance(pred, dict):
                rec_texts = pred.get("rec_texts", [])
                rec_scores = pred.get("rec_scores", [])
            rec_texts = rec_texts or []
            rec_scores = rec_scores or [0.6] * len(rec_texts)
            for txt, sc in zip(rec_texts, rec_scores):
                cleaned = (txt or "").strip()
                if cleaned:
                    text_parts.append(cleaned)
                    confs.append(float(sc))
    finally:
        try:
            os.remove(path)
        except OSError:
            pass
    if not text_parts:
        return "", 0.0
    joined = " ".join(text_parts).strip()
    return joined, float(sum(confs) / max(1, len(confs)))


def _normalize_token(text: str) -> str:
    out = (text or "").strip()
    out = re.sub(r"\s+", " ", out)
    out = out.replace("|", "")
    out = out.replace("O", "0")
    out = out.replace("l", "1")
    return out.strip()


def ocr_grid_cells(
    image: np.ndarray,
    x_lines: List[int],
    y_lines: List[int],
    pad: int = 10,
    debug_dir: Optional[str] = None,
) -> Tuple[List[List[str]], List[float]]:
    rows: List[List[str]] = []
    row_confidences: List[float] = []
    h, w = image.shape[:2]
    for r in range(len(y_lines) - 1):
        row_vals: List[str] = []
        confs: List[float] = []
        y1 = max(0, y_lines[r] - pad)
        y2 = min(h, y_lines[r + 1] + pad)
        for c in range(EXPECTED_COLS):
            x1 = max(0, x_lines[c] - pad)
            x2 = min(w, x_lines[c + 1] + pad)
            if (x2 - x1) < 20 or (y2 - y1) < 20:
                row_vals.append("")
                confs.append(0.0)
                continue
            crop = image[y1:y2, x1:x2]
            if debug_dir:
                os.makedirs(debug_dir, exist_ok=True)
                cv2.imwrite(os.path.join(debug_dir, f"cell_{r}_{c}.png"), crop)
            txt, conf = _ocr_crop_with_conf(crop)
            if conf < 0.5:
                retry_pad = pad + 6
                rx1 = max(0, x_lines[c] - retry_pad)
                rx2 = min(w, x_lines[c + 1] + retry_pad)
                ry1 = max(0, y_lines[r] - retry_pad)
                ry2 = min(h, y_lines[r + 1] + retry_pad)
                if (rx2 - rx1) < 20 or (ry2 - ry1) < 20:
                    row_vals.append(_normalize_token(txt))
                    confs.append(conf)
                    continue
                retry_crop = image[ry1:ry2, rx1:rx2]
                if debug_dir:
                    cv2.imwrite(os.path.join(debug_dir, f"cell_{r}_{c}_retry.png"), retry_crop)
                txt2, conf2 = _ocr_crop_with_conf(retry_crop)
                if conf2 > conf:
                    txt, conf = txt2, conf2
            row_vals.append(_normalize_token(txt))
            confs.append(conf)
        rows.append(row_vals)
        row_confidences.append(float(sum(confs) / max(1, len(confs))))
    return rows, row_confidences


def _extract_jte_rows(corrected_grid: List[List[str]]) -> List[List[str]]:
    if not corrected_grid:
        return [JTE_SCHEMA[:]]
    header_idx = -1
    for idx, row in enumerate(corrected_grid):
        joined = " ".join(cell.lower() for cell in row)
        if "sn" in joined and "name" in joined and "subject" in joined:
            header_idx = idx
            break
    if header_idx < 0:
        # Fallback: keep rows that look like class rows by serial number presence.
        out = [JTE_SCHEMA[:]]
        for row in corrected_grid:
            padded = (row + [""] * EXPECTED_COLS)[:EXPECTED_COLS]
            first = (padded[0] or "").strip()
            if not first:
                for cell in padded[:2]:
                    token = (cell or "").strip()
                    if token.isdigit() and 1 <= int(token) <= 8:
                        first = token
                        if token != padded[0]:
                            padded = [token] + padded[1:]
                        break
            if first.isdigit() and 1 <= int(first) <= 8:
                if not (padded[6] or "").strip() and (padded[7] or "").strip():
                    padded[6], padded[7] = padded[7], ""
                out.append(padded)
        while len(out) < 9:
            out.append([str(len(out)), "", "", "", "", "", "", "", ""])
        return out
    out = [JTE_SCHEMA[:]]
    for row in corrected_grid[header_idx + 1 :]:
        if not row:
            continue
        first = (row[0] or "").strip()
        if first.isdigit() and 1 <= int(first) <= 8:
            padded = (row + [""] * EXPECTED_COLS)[:EXPECTED_COLS]
            # Missing column recovery for subject/activity.
            if not (padded[6] or "").strip() and (padded[7] or "").strip():
                padded[6], padded[7] = padded[7], ""
            out.append(padded)
    while len(out) < 9:  # header + 8 data rows
        out.append([str(len(out)), "", "", "", "", "", "", "", ""])
    return out


def _clean_time(token: str) -> str:
    t = (token or "").replace(".", ":")
    t = re.sub(r"[^0-9:]", "", t)
    m = re.search(r"(\d{1,2}):?(\d{2})", t)
    if not m:
        return ""
    hh = int(m.group(1))
    mm = int(m.group(2))
    if 0 <= hh <= 23 and 0 <= mm <= 59:
        return f"{hh}:{mm:02d}"
    return ""


def _clean_class(token: str) -> str:
    raw = (token or "").lower()
    m = re.search(r"(\d{1,3})(st|nd|rd|th)?", raw)
    if not m:
        return ""
    n = int(m.group(1))
    if n > 12:
        n = int(str(n)[0])
    if n < 1 or n > 12:
        return ""
    if n == 1:
        return "1st"
    if n == 2:
        return "2nd"
    if n == 3:
        return "3rd"
    return f"{n}th"


def _fuzzy_subject(text: str) -> str:
    token = re.sub(r"[^a-z]", "", (text or "").lower())
    if not token:
        return ""
    subjects = ["Marathi", "GK", "Basic", "Maths", "English"]
    best = ""
    score = -1.0
    for sub in subjects:
        s = SequenceMatcher(None, token, sub.lower()).ratio()
        if s > score:
            score = s
            best = sub
    if token in {"ak", "ck", "gk"}:
        return "GK"
    return best if score >= 0.52 else ""


def _extract_name(text: str) -> str:
    # keep letters and spaces only, remove trailing time/class artifacts.
    t = re.sub(r"\d{1,2}[:.]\d{2}", " ", text or "")
    t = re.sub(r"\b\d+(st|nd|rd|th)?\b", " ", t, flags=re.IGNORECASE)
    t = re.sub(r"[^A-Za-z\s.]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _rebuild_jte_rows(rows: List[List[str]]) -> List[List[str]]:
    if not rows:
        return rows
    out = [rows[0][:]]
    serial = 1
    for row in rows[1:]:
        padded = (row + [""] * EXPECTED_COLS)[:EXPECTED_COLS]
        row_text = " ".join(padded)
        if not row_text.strip():
            continue

        # Extract times from entire row when columns are shifted.
        time_tokens = re.findall(r"\d{1,2}[:.]\d{2}", row_text)
        in_time = _clean_time(padded[2]) or (_clean_time(time_tokens[0]) if len(time_tokens) >= 1 else "")
        out_time = _clean_time(padded[3]) or (_clean_time(time_tokens[1]) if len(time_tokens) >= 2 else "")

        name = _extract_name(padded[1] or padded[0] or row_text)
        class_taught = _clean_class(padded[4] or row_text)

        # student count from dedicated col else from row tokens
        cnt = ""
        if re.fullmatch(r"\d{1,2}", (padded[5] or "").strip()):
            cnt = padded[5].strip()
        if not cnt:
            for n in re.findall(r"\b\d{1,2}\b", row_text):
                value = int(n)
                if 1 <= value <= 60:
                    if class_taught and value == int(re.sub(r"\D", "", class_taught)):
                        continue
                    cnt = str(value)
                    break

        subject = _fuzzy_subject(padded[6]) or _fuzzy_subject(padded[7]) or _fuzzy_subject(row_text)
        activity = (padded[7] or "").strip()
        if not activity and subject:
            # remove subject token from row text to keep residue as activity
            residue = re.sub(subject, "", row_text, flags=re.IGNORECASE)
            residue = _extract_name(residue)
            activity = residue if len(residue.split()) <= 4 else ""
        if not subject and activity:
            guess = _fuzzy_subject(activity)
            if guess:
                subject = guess
                activity = ""

        # skip near-empty rows
        payload = [name, in_time, out_time, class_taught, cnt, subject, activity]
        if sum(1 for x in payload if (x or "").strip()) < 2:
            continue

        out.append([
            str(serial),
            name,
            in_time,
            out_time,
            class_taught,
            cnt,
            subject,
            activity,
            (padded[8] or "").strip(),
        ])
        serial += 1
        if serial > 8:
            break

    while len(out) < 9:
        out.append([str(len(out)), "", "", "", "", "", "", "", ""])
    return out[:9]


def export_jte_excel(rows: List[List[str]], output_path: str) -> str:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "JTE"
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    thin = Side(style="thin", color="D9DDE5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for col_idx in range(1, EXPECTED_COLS + 1):
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[chr(64 + col_idx)].width = min(40, max(10, max_len + 2))

    wb.save(output_path)
    return output_path


def run_end_to_end_table_extraction(
    image_path: str,
    output_excel_path: Optional[str] = None,
    cfg: Optional[Dict[str, Any]] = None,
) -> Tuple[List[List[str]], Dict[str, Any], str]:
    """
    End-to-end OCR table extraction pipeline (strict order):
    preprocess -> detect grid -> extract cells -> validate/fix grid ->
    OCR per cell -> grid mapping -> correction engine -> validation/fail-safe -> JTE excel.
    """
    if output_excel_path is None:
        output_excel_path = os.path.join(os.path.dirname(__file__), "outputs", "jte_output.xlsx")

    image = cv2.imread(image_path)
    if image is None:
        raise ValueError(f"Image not found: {image_path}")

    debug_dir = os.path.join(os.path.dirname(__file__), "debug_cells")

    # 1) Preprocess (page-level, for contour detection)
    _ = preprocess_image(image)
    # 2) Table region isolation (CLASS DETAILS only)
    isolated = isolate_class_details_table(image)
    if not isolated:
        # hard fail-safe: grid detection failed
        normalized_grid = [JTE_SCHEMA[:]]
        excel_path = export_jte_excel(normalized_grid, output_excel_path)
        metadata = {
            "failsafe_triggered": True,
            "invalid_rows_ratio": 1.0,
            "grid_repair_failed": True,
            "pruned_count": 0,
            "candidate_list": [],
            "scoring_breakdown": [],
            "failsafe_reason": "grid_detection_failed",
        }
        return normalized_grid, metadata, excel_path
    table_img, table_bbox_abs = isolated

    # 2b) Detect grid on table only
    table_binary = preprocess_image(table_img)
    horizontal, vertical = detect_table_lines(table_binary)
    th, tw = table_img.shape[:2]
    table_bbox = (0, 0, tw - 1, th - 1)

    # 3) Cell extraction (intersection-based)
    points = _extract_intersections(horizontal, vertical)
    # 4) Grid validation + column alignment fix
    x_lines, y_lines, grid_meta = validate_and_fix_grid(points, table_bbox)
    # 5) OCR per cell (no full-image OCR)
    grid_rows, row_confidences = ocr_grid_cells(table_img, x_lines, y_lines, pad=10, debug_dir=debug_dir)
    # 6) Grid mapping is position-based by constructed grid rows/cols
    mapped_rows = [(row + [""] * EXPECTED_COLS)[:EXPECTED_COLS] for row in grid_rows]
    # 7-17) Correction, confidence arbitration, validation, fail-safe
    ml_predictor = _get_ml_predictor()
    correction_cfg = {"ENABLE_ML": bool(ml_predictor), "DEBUG": bool((cfg or {}).get("DEBUG", False))}
    corrected_grid, corr_meta = run_correction_engine(
        mapped_rows,
        schema=JTE_SCHEMA,
        row_confidences=row_confidences,
        ml_predictor=ml_predictor,
        cfg=correction_cfg,
    )

    invalid_ratio = float(corr_meta.get("invalid_rows_ratio", 0.0))
    failsafe = bool(corr_meta.get("failsafe_triggered", False)) or grid_meta.grid_repair_failed
    selected_grid = corrected_grid
    failsafe_reason = corr_meta.get("failsafe_reason")
    if failsafe and not failsafe_reason:
        failsafe_reason = "grid_repair_failed" if grid_meta.grid_repair_failed else "invalid_rows_ratio_exceeded"
    if failsafe:
        selected_grid = mapped_rows  # normalized/raw mapped grid fallback

    # 18) Export to JTE excel
    corrected_rows = _extract_jte_rows(selected_grid)
    corrected_rows = _rebuild_jte_rows(corrected_rows)
    excel_path = export_jte_excel(corrected_rows, output_excel_path)

    # Consolidated metadata
    pruned_count = 0
    candidate_list = []
    scoring_breakdown = []
    for row_meta in corr_meta.get("rows", []):
        for cell_meta in row_meta.get("cells", []):
            pruned_count += int(cell_meta.get("pruned_count", 0))
            candidate_list.append(cell_meta.get("candidates", []))
            scoring_breakdown.append(cell_meta.get("scoring_breakdown", []))

    metadata = {
        "failsafe_triggered": bool(failsafe),
        "failsafe_reason": failsafe_reason,
        "invalid_rows_ratio": invalid_ratio,
        "grid_repair_failed": bool(grid_meta.grid_repair_failed),
        "table_bbox": {
            "x1": int(table_bbox_abs[0]),
            "y1": int(table_bbox_abs[1]),
            "x2": int(table_bbox_abs[2]),
            "y2": int(table_bbox_abs[3]),
        },
        "debug_cells_dir": debug_dir,
        "pruned_count": pruned_count,
        "candidate_list": candidate_list,
        "scoring_breakdown": scoring_breakdown,
        "pipeline_trace": [
            "1_preprocess",
            "2_table_detection",
            "3_cell_extraction",
            "4_grid_validation_alignment_fix",
            "5_ocr_per_cell",
            "6_grid_mapping",
            "7_normalization",
            "8_context_correction",
            "9_deterministic_rules",
            "10_candidate_tracking",
            "11_candidate_pruning",
            "12_optional_ml",
            "13_confidence_scoring",
            "14_arbitration_tie_break",
            "15_confidence_floor",
            "16_final_validation",
            "17_fail_safe",
            "18_export_jte_excel",
        ],
    }
    return corrected_rows, metadata, excel_path
