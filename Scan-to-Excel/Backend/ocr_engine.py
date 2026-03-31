try:
    import cv2
    import pandas as pd
    import numpy as np
    import os
    import tempfile
    import re
    from paddlex import create_pipeline
    from correction_layer import run_correction_engine
    from ml_text_corrector import load_model
except ImportError as e:
    import sys
    print(f"\n[ERROR] Missing dependency: {e}")
    print(f"[ERROR] Current Python: {sys.version}")
    print("[ERROR] Please ensure you are using the correct virtual environment.")
    print("[ERROR] Run: source venv/bin/activate")
    print("[ERROR] Or select 'venv' as your interpreter in your IDE.\n")
    sys.exit(1)

# Disable oneDNN to fix Windows crash
os.environ["FLAGS_use_mkldnn"] = "0"
# Avoid slow external model host checks on every backend start.
os.environ["PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK"] = "True"

# Initialize PaddleX OCR pipeline once (expensive to load)
ocr = create_pipeline(pipeline="OCR")
_ML_PREDICTOR = None
_ML_ATTEMPTED = False


def get_ml_predictor():
    global _ML_PREDICTOR, _ML_ATTEMPTED
    if _ML_ATTEMPTED:
        return _ML_PREDICTOR

    model_path = os.path.join(os.path.dirname(__file__), "models", "ocr_text_corrector.pkl")
    _ML_PREDICTOR = load_model(model_path)
    _ML_ATTEMPTED = True
    if _ML_PREDICTOR:
        print(f"[INFO] Loaded OCR correction ML model: {model_path}")
    else:
        print("[INFO] OCR correction ML model not found; using deterministic-only correction.")
    return _ML_PREDICTOR


def infer_schema_from_table(table):
    """
    Infer a lightweight schema for correction layer.
    Prefer explicit class-details header if present, else widest row.
    """
    if not table:
        return []

    class_header = [
        "SN",
        "Volunteer/Teacher's Name",
        "In-time",
        "Out-time",
        "Class Taught",
        "No of students",
        "Subject",
        "Class Activity",
        "Homework",
    ]

    for row in table:
        row_text = " ".join((cell or "").lower() for cell in row)
        if "volunteer/teacher" in row_text and "subject" in row_text and "homework" in row_text:
            return class_header[: len(row)]

    widest = max(table, key=lambda r: len(r)) if table else []
    schema = []
    for idx, cell in enumerate(widest):
        token = normalize_whitespace(cell)
        schema.append(token if token else f"col_{idx}")
    return schema


def apply_correction_layer(table):
    """
    Run post-OCR correction engine. Falls back to original table on any failure.
    """
    if not table:
        return table

    schema = infer_schema_from_table(table)
    ml_predictor = get_ml_predictor()
    try:
        corrected, _meta = run_correction_engine(
            table,
            schema=schema,
            row_confidences=None,
            ml_predictor=ml_predictor,
            cfg={"ENABLE_ML": bool(ml_predictor), "DEBUG": False},
        )
        return corrected
    except Exception as exc:
        print(f"[WARN] Correction layer failed; using uncorrected table. Reason: {exc}")
        return table


# ─── STEP 1: PREPROCESS IMAGE ───────────────────────────────────────────────────

def preprocess_image(image):
    """Convert to grayscale for line detection."""
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    return gray


def detect_document_bbox(image, min_area_ratio=0.08):
    """
    Detect a bright document-like region (useful when screenshots include UI chrome).
    Returns (x1, y1, x2, y2) or None.
    """
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    img_h, img_w = gray.shape
    img_area = img_h * img_w

    _, bright_mask = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY)
    close_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (15, 15))
    bright_mask = cv2.morphologyEx(bright_mask, cv2.MORPH_CLOSE, close_kernel, iterations=1)

    contours_info = cv2.findContours(bright_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    contours = contours_info[0] if len(contours_info) == 2 else contours_info[1]
    if not contours:
        return None

    best_bbox = None
    best_score = -1.0

    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        area = w * h
        if area < img_area * min_area_ratio:
            continue
        if area > img_area * 0.98:
            continue

        contour_area = cv2.contourArea(contour)
        rectangularity = float(contour_area) / float(max(area, 1))
        if rectangularity < 0.45:
            continue

        aspect = float(w) / float(max(h, 1))
        if aspect < 0.4 or aspect > 3.5:
            continue

        score = (area / img_area) * 1.2 + rectangularity
        if score > best_score:
            best_score = score
            best_bbox = (x, y, x + w, y + h)

    return best_bbox


# ─── STEP 2: DETECT TABLE GRID LINES ────────────────────────────────────────────

def detect_table_lines(gray):
    """Detect horizontal and vertical lines to find the table grid."""
    # Binary threshold for line detection
    _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY_INV)

    img_h, img_w = thresh.shape

    # Horizontal lines
    h_kernel_len = max(img_w // 8, 40)
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (h_kernel_len, 1))
    horizontal = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)

    # Vertical lines
    v_kernel_len = max(img_h // 8, 40)
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, v_kernel_len))
    vertical = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, vertical_kernel, iterations=2)

    return horizontal, vertical


# ─── STEP 3: FIND ROW AND COLUMN BOUNDARIES ─────────────────────────────────────

def find_boundaries(line_mask, axis):
    """
    Find boundaries from a line mask.
    axis=1 for rows (sum across columns), axis=0 for columns (sum across rows).
    """
    sums = np.sum(line_mask, axis=axis)
    threshold = np.max(sums) * 0.3 if np.max(sums) > 0 else 0

    line_positions = np.where(sums > threshold)[0]

    if len(line_positions) == 0:
        return []

    # Cluster nearby positions into single boundaries
    boundaries = []
    cluster_start = line_positions[0]

    for i in range(1, len(line_positions)):
        if line_positions[i] - line_positions[i - 1] > 3:
            boundaries.append((cluster_start + line_positions[i - 1]) // 2)
            cluster_start = line_positions[i]

    boundaries.append((cluster_start + line_positions[-1]) // 2)

    return sorted(boundaries)


def detect_table_bbox(gray, horizontal, vertical, min_area_ratio=0.04):
    """
    Detect the primary table bounding box from line masks.
    Returns (x1, y1, x2, y2) or None.
    """
    combined = cv2.bitwise_or(horizontal, vertical)
    img_h, img_w = combined.shape
    img_area = img_h * img_w

    # Merge nearby grid lines into connected table blocks.
    kernel_w = max(20, img_w // 60)
    kernel_h = max(20, img_h // 60)
    merge_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (kernel_w, kernel_h))
    merged = cv2.dilate(combined, merge_kernel, iterations=2)

    contours_info = cv2.findContours(merged, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    contours = contours_info[0] if len(contours_info) == 2 else contours_info[1]
    if not contours:
        return None

    best_bbox = None
    best_score = -1.0

    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        area = w * h
        if area < img_area * min_area_ratio:
            continue
        if area > img_area * 0.9:
            continue

        roi = combined[y:y + h, x:x + w]
        density = float(np.count_nonzero(roi)) / float(max(area, 1))
        if density < 0.002:
            continue

        # Prefer paper-like bright table regions over dark UI grid regions.
        gray_roi = gray[y:y + h, x:x + w]
        brightness = float(np.mean(gray_roi)) / 255.0 if gray_roi.size else 0.0
        area_ratio = float(area) / float(max(img_area, 1))

        # Weighted score tuned for screenshot-heavy inputs:
        # density (grid strength) + brightness (paper) + moderate size.
        score = (density * 800.0) + (brightness * 3.0) + (area_ratio * 0.8)

        if score > best_score:
            best_score = score
            best_bbox = (x, y, x + w, y + h)

    return best_bbox


def filter_ocr_results_to_bbox(ocr_results, bbox, padding=8):
    """Keep only OCR boxes whose center lies inside table bbox (+ padding)."""
    if not bbox:
        return ocr_results

    x1, y1, x2, y2 = bbox
    x1 -= padding
    y1 -= padding
    x2 += padding
    y2 += padding

    filtered = []
    for text, cx, cy in ocr_results:
        if x1 <= cx <= x2 and y1 <= cy <= y2:
            filtered.append((text, cx, cy))
    return filtered


def is_daily_report_text(text):
    lowered = (text or "").lower()
    signals = [
        "daily centre report",
        "daily checklist",
        "class details",
        "thought of the day",
        "volunteer/teacher",
    ]
    return sum(1 for signal in signals if signal in lowered) >= 2


def table_from_ocr_results(ocr_results, row_threshold=15):
    """Group OCR detections into row-wise text by y proximity."""
    if not ocr_results:
        return []

    ordered = sorted(ocr_results, key=lambda r: (r[2], r[1]))
    rows = []
    current_row = []
    current_y = None

    for text, cx, cy in ordered:
        if current_y is None:
            current_y = cy

        if abs(cy - current_y) <= row_threshold:
            current_row.append((cx, text))
        else:
            rows.append(current_row)
            current_row = [(cx, text)]
            current_y = cy

    if current_row:
        rows.append(current_row)

    table = []
    for row in rows:
        row.sort(key=lambda item: item[0])
        table.append([cell[1] for cell in row])

    return table


# ─── STEP 4: OCR THE FULL IMAGE, THEN MAP TO GRID ───────────────────────────────

def ocr_full_image(image_path, use_enhanced_fallback=False):
    """
    Run PaddleOCR on the full image and return list of (text, center_x, center_y).
    This is MUCH more accurate than cropping tiny cells and OCR-ing each separately.
    """
    results = []

    def _collect(predictions):
        collected = []
        for pred in predictions:
            rec_texts = None
            dt_polys = None

            if hasattr(pred, "rec_texts"):
                rec_texts = pred.rec_texts
                dt_polys = pred.dt_polys
            elif isinstance(pred, dict):
                rec_texts = pred.get("rec_texts", [])
                dt_polys = pred.get("dt_polys", [])

            if rec_texts and dt_polys is not None:
                for text, poly in zip(rec_texts, dt_polys):
                    value = text.strip()
                    if not value:
                        continue
                    poly = np.array(poly)
                    cx = float(np.mean(poly[:, 0]))
                    cy = float(np.mean(poly[:, 1]))
                    collected.append((value, cx, cy))
        return collected

    try:
        # Pass 1: original image
        predictions = list(ocr.predict(image_path))
        results = _collect(predictions)

        # Pass 2 (optional): contrast-enhanced fallback for sparse text results.
        if use_enhanced_fallback and len(results) < 24:
            source = cv2.imread(image_path)
            if source is not None:
                gray = cv2.cvtColor(source, cv2.COLOR_BGR2GRAY)
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
                enhanced = clahe.apply(gray)
                enhanced = cv2.fastNlMeansDenoising(enhanced, h=9)
                enhanced = cv2.cvtColor(enhanced, cv2.COLOR_GRAY2BGR)

                with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                    cv2.imwrite(tmp.name, enhanced)
                    enhanced_path = tmp.name

                try:
                    fallback_predictions = list(ocr.predict(enhanced_path))
                    fallback_results = _collect(fallback_predictions)
                finally:
                    try:
                        os.remove(enhanced_path)
                    except OSError:
                        pass

                if fallback_results:
                    # Merge near-duplicates by coarse center buckets.
                    merged = {}
                    for text, cx, cy in results + fallback_results:
                        key = (int(cx // 8), int(cy // 8))
                        old = merged.get(key, "")
                        if len(text) > len(old):
                            merged[key] = text
                    results = [(text, key[0] * 8.0, key[1] * 8.0) for key, text in merged.items()]
    except Exception as e:
        import traceback
        print(f"[WARN] PaddleOCR predict failed: {e}")
        traceback.print_exc()

    return results



def assign_text_to_grid(ocr_results, row_bounds, col_bounds):
    """
    Map each OCR text result to the correct grid cell based on its center position.
    Returns a 2D list (rows × cols) of text strings.
    """
    num_rows = len(row_bounds) - 1
    num_cols = len(col_bounds) - 1

    # Initialize empty grid
    grid = [["" for _ in range(num_cols)] for _ in range(num_rows)]

    for text, cx, cy in ocr_results:
        if not text.strip():
            continue

        # Find which row this text belongs to
        row_idx = -1
        for r in range(num_rows):
            if row_bounds[r] <= cy <= row_bounds[r + 1]:
                row_idx = r
                break

        # Find which column this text belongs to
        col_idx = -1
        for c in range(num_cols):
            if col_bounds[c] <= cx <= col_bounds[c + 1]:
                col_idx = c
                break

        if row_idx >= 0 and col_idx >= 0:
            # Append text if cell already has content (rare)
            if grid[row_idx][col_idx]:
                grid[row_idx][col_idx] += " " + text
            else:
                grid[row_idx][col_idx] = text

    return grid


def fix_merged_cells(grid):
    """
    Fix cells where OCR merged adjacent column values.
    e.g. '2DS' in Species col with empty Plot col → Plot='2', Species='DS'
    e.g. '2 DM' in Species col with empty Plot col → Plot='2', Species='DM'
    """
    import re

    for row in grid:
        for col_idx in range(1, len(row)):
            cell = row[col_idx].strip()
            prev_cell = row[col_idx - 1].strip()

            # If previous cell is empty and this cell starts with a digit followed by text
            if prev_cell == "" and cell:
                # Pattern: "2DS" or "2 DS" or "2DM" or "2 DM"
                match = re.match(r'^(\d+)\s*([A-Za-z].*)$', cell)
                if match:
                    row[col_idx - 1] = match.group(1)
                    row[col_idx] = match.group(2)

    return grid


# ─── STEP 5: VALIDATE AND CLEAN ──────────────────────────────────────────────────

def validate_and_clean(table):
    """Clean OCR artifacts and remove empty rows."""
    if not table:
        return table

    cleaned = []

    for row in table:
        cleaned_row = []
        for cell in row:
            text = cell.strip()

            # Remove common OCR noise
            for char in ["|", "]", "[", "\\", "{", "}", "~", "`"]:
                text = text.replace(char, "")

            # Fix common OCR mistakes
            text = text.replace("l/", "1/")    # date fix
            text = text.replace("O/", "0/")    # date fix

            cleaned_row.append(text.strip())

        # Skip completely empty rows
        if any(cell.strip() for cell in cleaned_row):
            cleaned.append(cleaned_row)

    return cleaned


# ─── STEP 6: NORMALIZE COLUMNS ──────────────────────────────────────────────────

def normalize_columns(table):
    """Ensure all rows have the same number of columns."""
    if not table:
        return table

    max_cols = max(len(row) for row in table)

    normalized = []
    for row in table:
        while len(row) < max_cols:
            row.append("")
        normalized.append(row[:max_cols])

    return normalized


def is_numeric_like(value):
    token = normalize_whitespace(value)
    if not token:
        return False
    token = token.replace(",", "")
    return bool(re.fullmatch(r"-?\d+(\.\d+)?", token))


def repair_generic_table_structure(table):
    """
    Repair common OCR misalignment for merged-cell tables:
    - Fill down grouped first-column labels where source had rowspan/merged cells.
    - Absorb footnote-only fragments that appear as shifted mini-rows.
    """
    if not table:
        return table

    cols = max(len(row) for row in table)
    if cols < 4:
        return table

    repaired = [row[:] for row in table]

    # Step 0: Some scans create a dummy empty leading column; remove it when obvious.
    first_col_empty = sum(1 for row in repaired if not normalize_whitespace(row[0]))
    second_col_filled = sum(1 for row in repaired if len(row) > 1 and normalize_whitespace(row[1]))
    if cols >= 5 and first_col_empty >= int(0.7 * len(repaired)) and second_col_filled >= 3:
        repaired = [row[1:] for row in repaired]
        cols -= 1

    # Step 1: Fill down first-column group labels when first column is blank.
    current_group = ""
    for idx, row in enumerate(repaired):
        row += [""] * (cols - len(row))
        first = normalize_whitespace(row[0])
        second = normalize_whitespace(row[1]) if cols > 1 else ""

        if first and second and not is_numeric_like(second):
            current_group = first
            continue

        if not first and second and current_group and not is_numeric_like(second):
            # Do not fill obvious header rows.
            lowered = second.lower()
            if not any(keyword in lowered for keyword in ["table", "policy functions", "expenditure by"]):
                repaired[idx][0] = current_group

    # Step 1.5: Reconcile split numeric cells across adjacent rows.
    # Example:
    # prev: [Financial, 22.5, ""]
    # curr: [Information, "", "30.57 14.8"]
    # next: ["", 10.2, ""]
    # becomes:
    # prev last=30.57, curr numeric=10.2/14.8, next dropped.
    drop_rows = set()
    if cols >= 4:
        value_col_left = cols - 2
        value_col_right = cols - 1
        for idx in range(1, len(repaired)):
            row = repaired[idx]
            prev = repaired[idx - 1]

            right_text = normalize_whitespace(row[value_col_right])
            left_text = normalize_whitespace(row[value_col_left])
            right_numbers = re.findall(r"-?\d+(?:\.\d+)?", right_text)

            if len(right_numbers) >= 2 and not left_text:
                prev_left = normalize_whitespace(prev[value_col_left])
                prev_right = normalize_whitespace(prev[value_col_right])
                if is_numeric_like(prev_left) and not prev_right:
                    prev[value_col_right] = right_numbers[0]
                    row[value_col_right] = right_numbers[1]

                    # Pull left numeric value from next lightweight row if present.
                    if idx + 1 < len(repaired):
                        nxt = repaired[idx + 1]
                        nxt_left = normalize_whitespace(nxt[value_col_left])
                        nxt_right = normalize_whitespace(nxt[value_col_right])
                        non_empty_nxt = [normalize_whitespace(x) for x in nxt if normalize_whitespace(x)]
                        if is_numeric_like(nxt_left) and not nxt_right and len(non_empty_nxt) <= 2:
                            row[value_col_left] = nxt_left
                            drop_rows.add(idx + 1)

    # Step 2: Merge footnote-marker rows into previous row and drop them.
    for idx in range(1, len(repaired)):
        if idx in drop_rows:
            continue
        row = repaired[idx]
        non_empty_positions = [i for i, cell in enumerate(row) if normalize_whitespace(cell)]
        if len(non_empty_positions) == 0:
            continue

        non_empty_values = [normalize_whitespace(row[i]) for i in non_empty_positions]
        has_single_digit_fragment = any(re.fullmatch(r"\d{1,2}", val) for val in non_empty_values)
        has_numeric_value = any(is_numeric_like(val) for val in non_empty_values)

        # Row like: ["", "2", "", "30.57"] should be folded into previous logical row.
        if len(non_empty_positions) <= 2 and has_single_digit_fragment:
            prev = repaired[idx - 1]

            # Append footnote marker to nearest previous text cell.
            for val in non_empty_values:
                if re.fullmatch(r"\d{1,2}", val):
                    for col in range(min(2, cols - 1), -1, -1):
                        prev_text = normalize_whitespace(prev[col])
                        if prev_text and not is_numeric_like(prev_text):
                            prev[col] = f"{prev_text} {val}"
                            break

            # If this row carries a numeric tail and previous row lacks right-most value, shift it up.
            numeric_candidates = [val for val in non_empty_values if is_numeric_like(val)]
            if numeric_candidates:
                candidate = numeric_candidates[-1]
                for col in range(cols - 1, -1, -1):
                    if not normalize_whitespace(prev[col]):
                        prev[col] = candidate
                        break

            drop_rows.add(idx)

        # Row mostly empty with only a standalone footnote number should be removed.
        elif len(non_empty_positions) == 1 and re.fullmatch(r"\d{1,2}", non_empty_values[0]):
            drop_rows.add(idx)

    cleaned = [row for idx, row in enumerate(repaired) if idx not in drop_rows]
    return normalize_columns(cleaned)


def looks_like_daily_report(table):
    """Detect the recurring daily centre report layout."""
    flattened = " ".join(cell.lower() for row in table for cell in row if cell).strip()
    signals = [
        "thought of the day",
        "daily checklist",
        "class details",
        "centre",
        "volunteer",
    ]
    return sum(1 for signal in signals if signal in flattened) >= 3


def normalize_whitespace(text):
    return re.sub(r"\s+", " ", text or "").strip()


def extract_first_match(text, patterns):
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return normalize_whitespace(match.group(1))
    return ""


def infer_day_from_date(date_text):
    try:
        parsed = pd.to_datetime(date_text, dayfirst=True, errors="coerce")
        if pd.isna(parsed):
            return ""
        return parsed.day_name()[:3].upper()
    except Exception:
        return ""


def normalize_yes_no_token(token):
    cleaned = normalize_whitespace(token).upper().strip(":;,. ")
    yes_aliases = {"Y", "YES", "V", "T", "I", "L", "1", "7", ">", "S", "4"}
    # Common handwritten/checkmark OCR substitutions seen in checklist cells.
    yes_aliases.update({"E", "P", "H", "C", "O"})
    no_aliases = {"N", "NO", "X"}
    if cleaned in yes_aliases:
        return "Y"
    if cleaned in no_aliases:
        return "N"
    if cleaned and cleaned[0] in yes_aliases and not cleaned.startswith("N"):
        return "Y"
    return cleaned


def extract_yes_no_from_text(text):
    match = re.search(r"[:\-\s]([YyNnVvXx><7])\s*$", normalize_whitespace(text))
    if match:
        return normalize_yes_no_token(match.group(1))

    tokens = re.findall(r"\b([YyNnVvXx])\b", text or "")
    if tokens:
        return normalize_yes_no_token(tokens[-1])

    return ""


def clean_time_token(value):
    token = normalize_whitespace(value).replace(".", ":")
    token = re.sub(r"[^0-9:]", "", token)
    match = re.search(r"(\d{1,2}):?(\d{2})", token)
    if not match:
        return ""
    hour = int(match.group(1))
    minute = int(match.group(2))
    if hour > 23 or minute > 59:
        return ""
    return f"{hour}:{minute:02d}"


def clean_class_token(value):
    token = normalize_whitespace(value)
    if not token:
        return ""
    ordinal_match = re.search(r"(\d{1,3})(st|nd|rd|th)?", token, flags=re.IGNORECASE)
    if ordinal_match:
        number = ordinal_match.group(1)
        suffix = (ordinal_match.group(2) or "th").lower()
        if ordinal_match.group(2) is None and int(number) > 12:
            # OCR noise heuristic: use leading digit as class ordinal (e.g. 228 -> 2nd).
            lead_digit = number[0]
            n = int(lead_digit)
            if n == 1:
                return "1st"
            if n == 2:
                return "2nd"
            if n == 3:
                return "3rd"
            return f"{n}th"
        if number.endswith("1") and number != "11":
            suffix = "st"
        elif number.endswith("2") and number != "12":
            suffix = "nd"
        elif number.endswith("3") and number != "13":
            suffix = "rd"
        elif suffix not in {"st", "nd", "rd"}:
            suffix = "th"
        return f"{number}{suffix}"
    return token


def clean_numeric_token(value):
    digits = re.findall(r"\d+", value or "")
    return digits[0] if digits else ""


def clean_name_token(value):
    token = normalize_whitespace(value)
    token = re.sub(r"[^A-Za-z\s.]", "", token)
    return normalize_whitespace(token)


def clean_subject_token(value):
    from difflib import SequenceMatcher

    token = normalize_whitespace(value)
    token = re.sub(r"[^A-Za-z\s]", "", token)
    token = normalize_whitespace(token).title()
    lowered = token.lower()
    if lowered in {"maathi", "marati", "maati"}:
        return "Marathi"
    if lowered in {"mavathi", "mathi"}:
        return "Marathi"
    if lowered in {"gk", "ak", "ck"}:
        return "GK"
    # Fuzzy fallback for noisy spellings.
    candidates = ["Marathi", "GK", "Basic", "Maths", "English"]
    joined = re.sub(r"[^a-z]", "", lowered)
    if joined:
        best = ""
        score = -1.0
        for cand in candidates:
            s = SequenceMatcher(None, joined, cand.lower()).ratio()
            if s > score:
                score = s
                best = cand
        if score >= 0.55:
            return best
    return token


def clean_activity_token(value):
    token = normalize_whitespace(value)
    if not token:
        return ""
    token = token.replace("Wuiting", "Writing")
    token = token.replace("Readidg", "Reading")
    token = token.replace("Wreading", "Wo Reading")
    token = token.replace("classroom", "Classroom")
    token = re.sub(r"\s+", " ", token).strip()
    if token in {"a", "-11-", "--"}:
        return ""
    return token


def split_subject_and_activity(subject_text, activity_text):
    subject = clean_subject_token(subject_text)
    activity = clean_activity_token(activity_text)

    if subject and subject != clean_subject_token(""):
        # If subject cell contains both subject + activity, split by first token.
        raw = normalize_whitespace(subject_text)
        if raw and len(raw.split()) > 1:
            parts = raw.split()
            first = clean_subject_token(parts[0])
            if first in {"Marathi", "GK", "Basic", "Maths", "English"}:
                subject = first
                if not activity:
                    activity = clean_activity_token(" ".join(parts[1:]))

    if not subject and activity:
        # Sometimes subject is moved to activity column.
        first = clean_subject_token(activity.split()[0])
        if first in {"Marathi", "GK", "Basic", "Maths", "English"}:
            subject = first
            activity = clean_activity_token(" ".join(activity.split()[1:]))

    return subject, activity


def parse_compact_date_token(token):
    digits = re.sub(r"\D", "", token or "")
    if len(digits) == 6:
        day = digits[:2]
        month = digits[2:4]
        year = digits[4:6]
        if 1 <= int(month) <= 12:
            return f"{day}/{month}/{year}"
    if len(digits) == 8:
        day = digits[:2]
        year = digits[-2:]
        middle = digits[2:6]
        month_candidates = []
        for i in range(0, len(middle) - 1):
            m = middle[i:i + 2]
            if m.isdigit() and 1 <= int(m) <= 12:
                month_candidates.append(m)
        if month_candidates:
            month = sorted(month_candidates, key=lambda m: (not m.startswith("0"), int(m)))[0]
            return f"{day}/{month}/{year}"
    return ""


def infer_date_from_lines(lines):
    for line in lines[:5]:
        if "date" not in line.lower():
            continue
        compact = parse_compact_date_token(line)
        if compact:
            return compact
        raw = line.replace("|", "/").replace("\\", "/")
        m = re.search(r"(\d{1,2})\D+(\d{1,2})\D+(\d{2,4})", raw)
        if m:
            d = int(m.group(1))
            mo = int(m.group(2))
            y = m.group(3)[-2:]
            if 1 <= d <= 31 and 1 <= mo <= 12:
                return f"{d:02d}/{mo:02d}/{y}"
    return ""


def extract_checklist_value(section_text, keywords):
    lines = [normalize_whitespace(line) for line in section_text.splitlines() if normalize_whitespace(line)]

    for line in lines:
        lowered = line.lower()
        if all(keyword in lowered for keyword in keywords):
            direct = re.search(r"[:\-]\s*([A-Za-z0-9>]+)\s*$", line)
            if direct:
                return normalize_yes_no_token(direct.group(1))

    inline_pattern = r"{}[^A-Za-z0-9]{{0,8}}([A-Za-z0-9>])".format(r".*".join(keywords))
    inline = re.search(inline_pattern, section_text, flags=re.IGNORECASE | re.DOTALL)
    if inline:
        return normalize_yes_no_token(inline.group(1))

    return ""


def extract_checklist_values_from_rows(table):
    checklist_markers = [
        ("Centre started on time", ["centre", "start"]),
        ("Students wore I-Cards", ["students", "wore"]),
        ("Volunteers wore I-Cards", ["volunteer", "wore"]),
        ("Footwears placed properly", ["footwear"]),
        ("Prayer Conducted", ["prayer", "conduct"]),
        ("Explained the Thought", ["explained", "thought"]),
        ("Physical Activity", ["physical", "activity"]),
        ("Student's Attendance taken", ["student", "attendance"]),
        ("Closing prayer conducted", ["closing", "prayer"]),
        ("Centre closed on Time", ["centre", "closed"]),
    ]

    values_by_label = {label: "" for label, _ in checklist_markers}
    for row in table:
        row_cells = [normalize_whitespace(cell) for cell in row if normalize_whitespace(cell)]
        if not row_cells:
            continue
        row_text = " ".join(row_cells)
        lowered = row_text.lower()

        # Prefer per-cell matching so merged multi-checklist rows don't share one trailing token.
        for cell_text in row_cells:
            cell_lower = cell_text.lower()
            for label, keywords in checklist_markers:
                if all(keyword in cell_lower for keyword in keywords):
                    guess = extract_yes_no_from_text(cell_text)
                    if guess:
                        values_by_label[label] = guess

        # Fallback to row-level match only for labels still missing.
        for label, keywords in checklist_markers:
            if values_by_label[label]:
                continue
            if all(keyword in lowered for keyword in keywords):
                guess = extract_yes_no_from_text(row_text)
                if guess:
                    values_by_label[label] = guess

        # Some scans merge 2-3 checklist items in one cell; extract inline.
        for label, keywords in checklist_markers:
            if values_by_label[label]:
                continue
            inline_pattern = r"{}[^A-Za-z0-9]{{0,20}}([YyNnVvXx><7])".format(r".*".join(keywords))
            inline_match = re.search(inline_pattern, row_text, flags=re.IGNORECASE)
            if inline_match:
                values_by_label[label] = normalize_yes_no_token(inline_match.group(1))

    normalized_lines = []
    for label, _ in checklist_markers:
        value = normalize_yes_no_token(values_by_label[label])
        normalized_lines.append(f"{label}: {value}".rstrip())
    return normalized_lines


def parse_class_row_from_text(row_text):
    row_text = normalize_whitespace(row_text)
    if not row_text:
        return None

    serial_match = re.match(r"^(\d{1,2})\b", row_text)
    if not serial_match:
        return None

    serial = serial_match.group(1)
    text_after_serial = normalize_whitespace(row_text[len(serial):])

    time_matches = re.findall(r"\b(\d{1,2}[:.]\d{2})\b", text_after_serial)
    if len(time_matches) < 2:
        return None

    in_time = clean_time_token(time_matches[0])
    out_time = clean_time_token(time_matches[1])
    if not in_time or not out_time:
        return None

    first_time_index = text_after_serial.find(time_matches[0])
    second_time_index = text_after_serial.find(time_matches[1], first_time_index + len(time_matches[0]))

    teacher = clean_name_token(text_after_serial[:first_time_index])
    tail = normalize_whitespace(text_after_serial[second_time_index + len(time_matches[1]):])
    tail_tokens = tail.split()

    class_taught = ""
    no_of_students = ""
    subject = ""
    activity = ""
    homework = ""

    if tail_tokens:
        class_taught = clean_class_token(tail_tokens[0])
    if len(tail_tokens) > 1:
        no_of_students = clean_numeric_token(tail_tokens[1])
    if len(tail_tokens) > 2:
        subject = clean_subject_token(tail_tokens[2])
    if len(tail_tokens) > 3:
        activity = normalize_whitespace(" ".join(tail_tokens[3:]))

    return [serial, teacher, in_time, out_time, class_taught, no_of_students, subject, activity, homework]


def extract_class_rows(table):
    class_rows = []
    orphan_lines = []
    in_class_section = False

    for row in table:
        normalized_row = [normalize_whitespace(cell) for cell in row]
        joined = " ".join(cell.lower() for cell in normalized_row if cell)

        if "class details" in joined:
            in_class_section = True
            continue

        if not in_class_section:
            continue

        if "any other extra activities" in joined or "visitors information" in joined:
            break

        parsed_from_line = parse_class_row_from_text(" ".join(normalized_row))
        if parsed_from_line:
            class_rows.append(parsed_from_line)
            continue

        if normalized_row and re.fullmatch(r"\d+", normalized_row[0] or ""):
            padded = normalized_row + [""] * (9 - len(normalized_row))
            serial = padded[0]
            teacher = clean_name_token(padded[1])
            in_time = clean_time_token(padded[2]) or clean_time_token(" ".join(padded[2:4])) or padded[2]
            out_time = clean_time_token(padded[3]) or padded[3]

            remaining = [cell for cell in padded[4:] if cell]
            class_taught = ""
            no_of_students = ""
            subject = ""
            activity = ""
            homework = ""

            if remaining:
                if len(remaining) >= 4 and re.search(r"\d", remaining[1]):
                    class_taught = clean_class_token(remaining[0])
                    no_of_students = clean_numeric_token(remaining[1])
                    subject = clean_subject_token(remaining[2])
                    activity = normalize_whitespace(remaining[3])
                    homework = remaining[4] if len(remaining) > 4 else ""
                else:
                    no_of_students = clean_numeric_token(remaining[0])
                    subject = clean_subject_token(remaining[1]) if len(remaining) > 1 else ""
                    activity = normalize_whitespace(remaining[2]) if len(remaining) > 2 else ""
                    homework = remaining[3] if len(remaining) > 3 else ""

            # Handle common OCR shift: class cell holds student count, and subject+activity spill.
            if not no_of_students and re.fullmatch(r"\d+", normalize_whitespace(class_taught or "")):
                candidate_count = clean_numeric_token(class_taught)
                if candidate_count:
                    no_of_students = candidate_count
                    class_taught = ""

            merged_subject_activity = normalize_whitespace(padded[6])
            if merged_subject_activity and not activity:
                pieces = merged_subject_activity.split()
                if len(pieces) >= 2:
                    subject = clean_subject_token(pieces[0])
                    activity = normalize_whitespace(" ".join(pieces[1:]))

            if subject and subject.lower() in {"wreading", "reading", "wuiting", "writing"} and not activity:
                activity = subject
                if normalize_whitespace(padded[5]):
                    subject = clean_subject_token(padded[5])
                else:
                    subject = ""

            subject, activity = split_subject_and_activity(subject, activity)
            activity = clean_activity_token(activity)

            class_rows.append([
                serial,
                teacher,
                in_time,
                out_time,
                class_taught,
                no_of_students,
                subject,
                activity,
                homework,
            ])
            continue

        # Capture probable continuation rows where subject/activity is printed
        # in the next line under class table.
        non_empty = [cell for cell in normalized_row if cell]
        if non_empty:
            row_text = normalize_whitespace(" ".join(non_empty))
            lowered = row_text.lower()
            looks_like_header = any(
                key in lowered
                for key in [
                    "class details",
                    "volunteer/teacher",
                    "in-time",
                    "out-time",
                    "no of",
                    "taught",
                    "students",
                    "subject",
                    "class activity",
                    "homework",
                    "sn",
                ]
            )
            looks_like_main_row = bool(re.search(r"\b\d{1,2}\s*[:.]\s*\d{2}\b", row_text))
            starts_with_serial = bool(non_empty and re.fullmatch(r"\d+", non_empty[0]))
            has_alpha = bool(re.search(r"[A-Za-z]", row_text))
            if (
                not looks_like_header
                and not looks_like_main_row
                and not starts_with_serial
                and has_alpha
                and len(row_text) <= 40
            ):
                orphan_lines.append(row_text)

    class_rows = class_rows[:8]

    # Attach orphan continuation lines (subject/activity spillover) to rows in order.
    if class_rows and orphan_lines:
        target_idx = 0
        for orphan in orphan_lines:
            if target_idx >= len(class_rows):
                break
            tokens = orphan.split()
            subj_guess = clean_subject_token(tokens[0]) if tokens else ""
            activity_guess = normalize_whitespace(" ".join(tokens[1:])) if len(tokens) > 1 else ""
            activity_guess = clean_activity_token(activity_guess)

            row = class_rows[target_idx]
            if not normalize_whitespace(row[6]) and subj_guess in {"Marathi", "GK", "Basic", "Maths", "English"}:
                row[6] = subj_guess
            elif not normalize_whitespace(row[7]) and subj_guess:
                # If subject already present, spill this token into activity.
                row[7] = clean_activity_token(subj_guess)

            if activity_guess:
                existing = normalize_whitespace(row[7])
                merged = normalize_whitespace(f"{existing} {activity_guess}") if existing else activity_guess
                row[7] = clean_activity_token(merged)

            target_idx += 1

    # Remove lightweight mis-split rows (e.g. "6 | th" with no real payload).
    filtered_rows = []
    for row in class_rows:
        teacher = normalize_whitespace(row[1]).lower()
        class_taught = clean_class_token(row[4])
        subject = clean_subject_token(row[6])
        activity = normalize_whitespace(row[7])
        payload_score = sum(
            1
            for token in [teacher, row[2], row[3], class_taught, row[5], subject, activity]
            if normalize_whitespace(token)
        )
        if teacher in {"th", "st", "nd", "rd"} and payload_score <= 3:
            continue
        row[4] = class_taught
        row[6] = subject
        filtered_rows.append(row)
    class_rows = filtered_rows[:8]

    # Normalize row timings using the most common in/out time where OCR is noisy.
    if class_rows:
        from collections import Counter

        def _parse_minutes(token):
            match = re.match(r"^(\d{1,2}):(\d{2})$", normalize_whitespace(token))
            if not match:
                return None
            hour = int(match.group(1))
            minute = int(match.group(2))
            if hour > 23 or minute > 59:
                return None
            return hour * 60 + minute

        valid_in = [row[2] for row in class_rows if _parse_minutes(row[2]) is not None]
        valid_out = [row[3] for row in class_rows if _parse_minutes(row[3]) is not None]
        mode_in = Counter(valid_in).most_common(1)[0][0] if valid_in else ""
        mode_out = Counter(valid_out).most_common(1)[0][0] if valid_out else ""

        for row in class_rows:
            in_minutes = _parse_minutes(row[2])
            out_minutes = _parse_minutes(row[3])
            has_payload = any(normalize_whitespace(cell) for cell in row[1:])
            if not has_payload:
                continue

            if in_minutes is None and mode_in:
                row[2] = mode_in
                in_minutes = _parse_minutes(row[2])
            if out_minutes is None and mode_out:
                row[3] = mode_out
                out_minutes = _parse_minutes(row[3])

            # If out-time is implausibly earlier than in-time, snap to common out-time.
            if in_minutes is not None and out_minutes is not None and out_minutes < in_minutes and mode_out:
                row[3] = mode_out

    # Normalize serial numbers to 1..8 for consistent template output.
    for idx, row in enumerate(class_rows, start=1):
        row[0] = str(idx)
        row[6], row[7] = split_subject_and_activity(row[6], row[7])
        row[7] = clean_activity_token(row[7])
        if row[6] and row[6].lower() in {"wreading", "reading", "writing"}:
            if not row[7]:
                row[7] = clean_activity_token(row[6])
            row[6] = ""

    # Fill missing subject from dominant subject if row has activity but empty subject.
    subject_values = [normalize_whitespace(r[6]) for r in class_rows if normalize_whitespace(r[6])]
    mode_subject = ""
    if subject_values:
        from collections import Counter

        mode_subject = Counter(subject_values).most_common(1)[0][0]
    if mode_subject:
        for row in class_rows:
            if not normalize_whitespace(row[6]) and normalize_whitespace(row[7]):
                row[6] = mode_subject

    return class_rows


def reconstruct_daily_report(table):
    """Reshape OCR output into the expected daily report sheet."""
    flattened_lines = [normalize_whitespace(" ".join(cell for cell in row if cell)) for row in table]
    flattened_lines = [line for line in flattened_lines if line]
    flattened_text = "\n".join(flattened_lines)

    date_value = extract_first_match(flattened_text, [
        r"date[:\s]*([0-9]{1,2}[\/\-][0-9]{1,2}[\/\-][0-9]{2,4})",
        r"\b([0-9]{1,2}[\/\-][0-9]{1,2}[\/\-][0-9]{2,4})\b",
    ])
    if not date_value:
        compact_date = extract_first_match(flattened_text, [r"date[:\s]*([0-9]{6,8})"])
        date_value = parse_compact_date_token(compact_date)
    if not date_value:
        date_value = infer_date_from_lines(flattened_lines)
    total_students = ""
    for line in flattened_lines:
        line_l = line.lower()
        if "total" in line_l and "students" in line_l:
            nums = re.findall(r"\b(\d{1,3})\b", line)
            if nums:
                total_students = nums[-1]
                break
    if not total_students:
        total_students = extract_first_match(flattened_text, [
            r"total(?:\s+number)?(?:\s+students)?(?:\s+present)?[:\s]*([0-9]+)",
            r"students\s+present[:\s]*([0-9]+)",
        ])

    thought = extract_first_match(flattened_text, [
        r"thought of the day[:\s]*(.+?)(?:\n|daily checklist|class details|$)",
    ])
    if "daily checklist" in thought.lower():
        thought = ""

    checklist_values = extract_checklist_values_from_rows(table)
    if not any(line.endswith(": Y") or line.endswith(": N") for line in checklist_values):
        # Fallback to prior flattened extraction if row-aware extraction fails.
        fallback_markers = [
            ("Centre started on time", ["centre", "start"]),
            ("Students wore I-Cards", ["students", "wore"]),
            ("Volunteers wore I-Cards", ["volunteer", "wore"]),
            ("Footwears placed properly", ["footwear"]),
            ("Prayer Conducted", ["prayer", "conduct"]),
            ("Explained the Thought", ["explained", "thought"]),
            ("Physical Activity", ["physical", "activity"]),
            ("Student's Attendance taken", ["student", "attendance"]),
            ("Closing prayer conducted", ["closing", "prayer"]),
            ("Centre closed on Time", ["centre", "closed"]),
        ]
        checklist_values = []
        for label, keywords in fallback_markers:
            value = extract_checklist_value(flattened_text, keywords)
            checklist_values.append(f"{label}: {value}".rstrip())

    class_rows = extract_class_rows(table)
    inferred_students = 0
    inferred_count_cells = 0
    inferred_max = 0
    for row in class_rows:
        val = clean_numeric_token(row[5])
        if val:
            n = int(val)
            inferred_students += n
            inferred_count_cells += 1
            inferred_max = max(inferred_max, n)
    if inferred_students > 0:
        try:
            parsed_total = int(total_students) if total_students else 0
        except Exception:
            parsed_total = 0
        # Only use inferred total when header total is missing/clearly implausible
        # and inferred counts look sane.
        if (
            (parsed_total <= 0 or parsed_total > 80)
            and inferred_count_cells >= 2
            and inferred_max <= 25
        ):
            total_students = str(inferred_students)
    while len(class_rows) < 8:
        class_rows.append([str(len(class_rows) + 1), "", "", "", "", "", "", "", ""])

    reconstructed = [
        [
            f"DATE: {date_value}" if date_value else "DATE:",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            f"TOTAL NUMBER STUDENTS PRESENT: {total_students}" if total_students else "TOTAL NUMBER STUDENTS PRESENT:",
        ],
        ["THOUGHT OF THE DAY:", thought, "", "", "", "", "", "", ""],
        ["", "[Put Y for Yes N for No for the points mentioned below]", "", "", "", "", "", "", ""],
        checklist_values[:5] + ["", "", "", ""],
        checklist_values[5:] + ["", "", "", ""],
        ["SN", "Volunteer/Teacher's Name", "In-time", "Out-time", "Class Taught", "No of students", "Subject", "Class Activity", "Homework"],
    ]

    reconstructed.extend(class_rows)
    reconstructed.extend([
        ["ANY OTHER EXTRA ACTIVITIES:", "", "", "", "", "", "", "", ""],
        ["VISITORS INFORMATION ALONG WITH CONTACT DETAILS:", "", "", "", "", "", "", "", ""],
        ["ANY SUGGESTION/IDEA FOR FURTHER BETTERMENT OR CHALLENGES FACED:", "", "", "", "", "", "", "", ""],
    ])

    return reconstructed


# ─── FALLBACK: FULL-IMAGE OCR WITHOUT GRID ───────────────────────────────────────

def fallback_full_image_ocr(image_path, table_bbox=None):
    """
    Fallback when grid lines aren't detected.
    Use PaddleOCR on the full image and group by Y-position into rows.
    """
    ocr_results = ocr_full_image(image_path)
    ocr_results = filter_ocr_results_to_bbox(ocr_results, table_bbox)

    return table_from_ocr_results(ocr_results, row_threshold=15)


# ─── MAIN PIPELINE ───────────────────────────────────────────────────────────────

def process_image(image_path):
    """
    Full pipeline:
    Scanned document → preprocess → detect table structure →
    segment rows/columns/cells → OCR full image → map to grid →
    validate → reconstruct
    """
    temp_path = None
    try:
        # Load image
        image = cv2.imread(image_path)

        if image is None:
            raise ValueError(f"Image not found: {image_path}")

        working_image = image
        working_path = image_path

        # If the input is a full app screenshot, crop to the bright document block first.
        doc_bbox = detect_document_bbox(image)
        if doc_bbox:
            dx1, dy1, dx2, dy2 = doc_bbox
            doc_area = (dx2 - dx1) * (dy2 - dy1)
            img_area = image.shape[0] * image.shape[1]
            if doc_area < img_area * 0.95:
                working_image = image[dy1:dy2, dx1:dx2].copy()
                with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                    cv2.imwrite(tmp.name, working_image)
                    temp_path = tmp.name
                    working_path = temp_path

        img_h, img_w = working_image.shape[:2]
        max_dim = max(img_h, img_w)

        # Speed optimization: downscale very large images before OCR.
        if max_dim > 1900:
            scale = 1900.0 / float(max_dim)
            new_w = max(1, int(img_w * scale))
            new_h = max(1, int(img_h * scale))
            working_image = cv2.resize(working_image, (new_w, new_h), interpolation=cv2.INTER_AREA)
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                cv2.imwrite(tmp.name, working_image)
                temp_path = tmp.name
                working_path = temp_path

        # Step 1: Preprocess
        gray = preprocess_image(working_image)

        # Step 2: Detect table grid lines
        horizontal, vertical = detect_table_lines(gray)

        # Step 3: Detect main table region and find row/column boundaries inside it
        table_bbox = detect_table_bbox(gray, horizontal, vertical)
        if table_bbox:
            x1, y1, x2, y2 = table_bbox
            horizontal_roi = horizontal[y1:y2, x1:x2]
            vertical_roi = vertical[y1:y2, x1:x2]
            row_bounds = [y1 + v for v in find_boundaries(horizontal_roi, axis=1)]
            col_bounds = [x1 + v for v in find_boundaries(vertical_roi, axis=0)]
        else:
            row_bounds = find_boundaries(horizontal, axis=1)
            col_bounds = find_boundaries(vertical, axis=0)

        # Step 4: OCR the full image (much better than per-cell OCR)
        ocr_results_all = ocr_full_image(working_path, use_enhanced_fallback=True)

        # Daily report forms are more reliable with full-page OCR + template reconstruction.
        full_text = " ".join(text for text, _, _ in ocr_results_all)
        if is_daily_report_text(full_text):
            table = table_from_ocr_results(ocr_results_all, row_threshold=16)
            table = validate_and_clean(table)
            table = normalize_columns(table)
            table = reconstruct_daily_report(table)
            return table

        ocr_results = ocr_results_all
        ocr_results = filter_ocr_results_to_bbox(ocr_results, table_bbox)

        if len(row_bounds) >= 3 and len(col_bounds) >= 3 and ocr_results:
            # Map OCR results to grid cells
            table = assign_text_to_grid(ocr_results, row_bounds, col_bounds)
            # Fix merged cells (e.g. "2DM" → "2" + "DM")
            table = fix_merged_cells(table)
        elif ocr_results:
            # Fallback: group OCR results by Y-position
            print("[INFO] Grid lines not fully detected, grouping by position...")
            table = fallback_full_image_ocr(working_path, table_bbox=table_bbox)
        else:
            table = []

        # Step 5: Validate and clean
        table = validate_and_clean(table)

        # Step 6: Normalize columns
        table = normalize_columns(table)

        if looks_like_daily_report(table):
            table = reconstruct_daily_report(table)
            return table
        else:
            table = repair_generic_table_structure(table)

        return apply_correction_layer(table)
    finally:
        if temp_path:
            try:
                os.remove(temp_path)
            except OSError:
                pass


def process_pdf(pdf_path, max_pages=2, zoom=1.35):
    """
    Convert each PDF page to an image and extract table data page by page.
    Returns one merged table so the current frontend/download flow stays unchanged.
    """
    try:
        import fitz  # PyMuPDF
    except ImportError as e:
        raise ImportError(
            "PDF support requires PyMuPDF. Install with: pip install pymupdf"
        ) from e

    merged_table = []
    temp_files = []

    try:
        with fitz.open(pdf_path) as document:
            total_pages = min(len(document), max_pages)

            if total_pages == 0:
                return []

            for page_index in range(total_pages):
                page = document.load_page(page_index)
                matrix = fitz.Matrix(zoom, zoom)
                pix = page.get_pixmap(matrix=matrix, alpha=False)

                with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                    tmp.write(pix.tobytes("png"))
                    temp_path = tmp.name

                temp_files.append(temp_path)
                page_table = process_image(temp_path)

                if not page_table:
                    continue

                if page_index > 0 and merged_table:
                    merged_table.append([f"--- PAGE {page_index + 1} ---"])

                merged_table.extend(page_table)

        return normalize_columns(merged_table)
    finally:
        for temp_path in temp_files:
            try:
                os.remove(temp_path)
            except OSError:
                pass


# ─── SAVE TO EXCEL ───────────────────────────────────────────────────────────────

def save_to_excel(table, output_path):
    """Save the validated table to an Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Extracted Data"

    for row_index, row in enumerate(table, start=1):
        for col_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)

    thin_gray = Side(style="thin", color="D9DDE5")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    heading_fill = PatternFill(fill_type="solid", fgColor="F7F9FC")

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.row in {3, 7, 8}:
                cell.font = Font(bold=True)
                cell.fill = heading_fill

    column_widths = {
        "A": 24,
        "B": 18,
        "C": 13,
        "D": 13,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 12,
        "I": 14,
    }
    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width

    for row_number in range(1, len(table) + 1):
        sheet.row_dimensions[row_number].height = 34

    workbook.save(output_path)
